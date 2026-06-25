from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def extract_xlsx_hyperlinks(path: Path) -> list[str]:
    """Return embedded hyperlink targets from an XLSX workbook.

    read_only=False is required: openpyxl does not expose hyperlink attributes
    on ReadOnlyCell/EmptyCell objects.
    """
    wb = load_workbook(path, read_only=False, data_only=True)
    targets: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    hyperlink = getattr(cell, "hyperlink", None)
                    if hyperlink and hyperlink.target:
                        targets.append(hyperlink.target)
    finally:
        wb.close()
    return targets


class XlsxExtractor:
    def extract_text(self, path: Path) -> str:
        # ``read_only=True`` is more memory friendly, but openpyxl does not load
        # hyperlinks in that mode - the cells become ReadOnlyCell/EmptyCell
        # objects without a ``hyperlink`` attribute, so accessing it raised
        # AttributeError and aborted extraction. We need the hyperlink targets
        # for the link-discovery pipeline, so load the workbook normally and
        # keep ``data_only`` to read computed values rather than formulas.
        wb = load_workbook(path, read_only=False, data_only=True)
        parts: list[str] = []
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            parts.append(str(cell.value))
                        hyperlink = getattr(cell, "hyperlink", None)
                        if hyperlink and hyperlink.target:
                            parts.append(hyperlink.target)
        finally:
            wb.close()
        return "\n".join(parts)
